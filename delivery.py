"""
delivery.py - 搬入日自動計算 & HTML生成
- daily_omura_2.xlsx の「カレンダー」シートで稼働日を判定
- topシートの最新日付を注文日として使用
- 中2日後の稼働日を搬入日として計算
- レトロCRTデザイン1行HTMLを生成
"""

import datetime
import os
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None
    print("ERROR: openpyxl not installed.")

# ============================================================
# 設定
# ============================================================
EXCEL_FILE  = Path(r"C:\Users\moto8\OneDrive\デスクトップ\daily_omura_2.xlsx")
OUTPUT_HTML = Path(__file__).parent / "delivery.html"
SKIP_DAYS   = 2  # 中N日

WEEKDAY_EN = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]
MONTH_EN   = ["JAN","FEB","MAR","APR","MAY","JUN",
               "JUL","AUG","SEP","OCT","NOV","DEC"]


# ============================================================
# カレンダーシートから稼働日セットを読み込む
# ============================================================
def load_working_days(excel_path: Path) -> set:
    """
    カレンダーシートの稼働フラグ=1の日付をセットで返す
    """
    working = set()
    if openpyxl is None or not excel_path.exists():
        print(f"  WARN: {excel_path.name} not found. Using weekdays only.")
        return working

    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    if "カレンダー" not in wb.sheetnames:
        print("  WARN: カレンダーシートが見つかりません。")
        return working

    ws = wb["カレンダー"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_val, _, flag = row[0], row[1], row[2]
        if not date_val or flag is None:
            continue
        if isinstance(date_val, datetime.datetime):
            d = date_val.date()
        elif isinstance(date_val, datetime.date):
            d = date_val
        else:
            continue
        if int(flag) == 1:
            working.add(d)

    print(f"  Working days loaded: {len(working)} days")
    return working


# ============================================================
# topシートから最新注文日を取得
# ============================================================
def load_latest_order_date(excel_path: Path) -> datetime.date | None:
    """
    topシートのヘッダー行（1行目）の5列目以降から最新日付を返す
    """
    if openpyxl is None or not excel_path.exists():
        print(f"  WARN: {excel_path.name} not found.")
        return None

    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    if "top" not in wb.sheetnames:
        print("  WARN: topシートが見つかりません。")
        return None

    ws = wb["top"]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))

    latest = None
    for cell in header[4:]:  # 5列目以降が日付
        if cell is None:
            continue
        if isinstance(cell, datetime.datetime):
            d = cell.date()
        elif isinstance(cell, datetime.date):
            d = cell
        else:
            try:
                d = datetime.date.fromisoformat(str(cell).strip()[:10])
            except ValueError:
                continue
        if latest is None or d > latest:
            latest = d

    return latest


# ============================================================
# 搬入日計算（中N日後の稼働日）
# カレンダーにない日付はフォールバックで平日判定
# ============================================================
def calc_delivery_date(
    order_date: datetime.date,
    working_days: set,
    skip: int = 2
) -> datetime.date:

    def is_working(d: datetime.date) -> bool:
        if working_days:
            return d in working_days
        # フォールバック: 土日除外のみ
        return d.weekday() < 5

    d = order_date
    count = 0
    while count < skip:
        d += datetime.timedelta(days=1)
        if is_working(d):
            count += 1

    d += datetime.timedelta(days=1)
    while not is_working(d):
        d += datetime.timedelta(days=1)

    return d


# ============================================================
# HTML生成（CRT 1行表示）
# ============================================================
def generate_html(delivery: datetime.date, now: datetime.datetime) -> str:
    wd  = WEEKDAY_EN[delivery.weekday()]
    mon = MONTH_EN[delivery.month - 1]
    delivery_str = f"{wd}, {mon} {delivery.day}"

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Delivery Schedule</title>
  <style>
    @import url('https://fonts.googleapis.com/css2?family=Share+Tech+Mono&display=swap');
    *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{
      background: #020c02;
      display: flex;
      align-items: center;
      justify-content: center;
      min-height: 100vh;
      padding: 12px;
    }}
    .crt {{
      font-family: 'Share Tech Mono', 'Courier New', monospace;
      background: #020c02;
      border: 1px solid #005520;
      border-radius: 6px;
      overflow: hidden;
      position: relative;
      box-shadow: 0 0 18px rgba(0,255,65,0.06);
      width: 100%;
      max-width: 480px;
    }}
    .crt::before {{
      content: '';
      position: absolute; inset: 0;
      background: repeating-linear-gradient(
        0deg, transparent, transparent 2px,
        rgba(0,0,0,0.08) 2px, rgba(0,0,0,0.08) 4px
      );
      pointer-events: none;
      z-index: 10;
    }}
    .row {{
      display: flex;
      align-items: center;
      gap: 12px;
      padding: 13px 20px;
    }}
    .lbl {{
      font-size: 12px;
      color: #007a1f;
      letter-spacing: 1px;
    }}
    .val {{
      font-size: 15px;
      color: #00ff41;
      font-weight: bold;
      text-shadow: 0 0 6px rgba(0,255,65,0.4);
    }}
    .sep {{
      font-size: 13px;
      color: #005510;
    }}
    @keyframes blink {{ 0%,100%{{opacity:1}} 50%{{opacity:0}} }}
    .cursor {{
      display: inline-block;
      width: 8px; height: 14px;
      background: #00ff41;
      vertical-align: middle;
      margin-left: 3px;
      animation: blink 1s step-end infinite;
    }}
  </style>
</head>
<body>
  <div class="crt">
    <div class="row">
      <span class="lbl">TIME</span>
      <span class="val" id="clock">--:--</span>
      <span class="sep">&#9658;</span>
      <span class="lbl">DELIVERY</span>
      <span class="val">{delivery_str}<span class="cursor"></span></span>
    </div>
  </div>
  <script>
    function tick() {{
      const now = new Date();
      const hh = String(now.getHours()).padStart(2, '0');
      const mm = String(now.getMinutes()).padStart(2, '0');
      document.getElementById('clock').textContent = hh + ':' + mm;
    }}
    tick();
    setInterval(tick, 1000);
  </script>
</body>
</html>
"""


# ============================================================
# メイン
# ============================================================
def main():
    now = datetime.datetime.now()
    print(f"[{now:%Y/%m/%d %H:%M:%S}] delivery.py started")

    # カレンダーから稼働日を読み込み
    working_days = load_working_days(EXCEL_FILE)

    # topシートから最新注文日を取得
    order_date = load_latest_order_date(EXCEL_FILE)
    if order_date is None:
        order_date = now.date()
        print(f"  Fallback: using today as order date ({order_date})")
    else:
        print(f"  Order date (latest): {order_date}")

    # 搬入日計算
    delivery = calc_delivery_date(order_date, working_days, SKIP_DAYS)
    wd  = WEEKDAY_EN[delivery.weekday()]
    mon = MONTH_EN[delivery.month - 1]
    print(f"  Delivery date: {delivery} ({wd}, {mon} {delivery.day})")

    # HTML生成
    html = generate_html(delivery, now)
    OUTPUT_HTML.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_HTML, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  Generated: {OUTPUT_HTML}")

    # GitHub push
    import subprocess
    repo_dir = OUTPUT_HTML.parent
    try:
        subprocess.run(["git", "-C", str(repo_dir), "add", "delivery.html"], check=True)
        result = subprocess.run(
            ["git", "-C", str(repo_dir), "status", "--porcelain", "delivery.html"],
            capture_output=True, text=True
        )
        if result.stdout.strip():
            subprocess.run(
                ["git", "-C", str(repo_dir), "commit", "-m",
                 f"delivery: {now.strftime('%Y-%m-%d %H:%M')}"],
                check=True
            )
            subprocess.run(["git", "-C", str(repo_dir), "push", "origin", "main"], check=True)
            print("  GitHub push: OK")
        else:
            print("  GitHub push: skipped (no changes)")
    except subprocess.CalledProcessError as e:
        print(f"  GitHub push ERROR: {e}")

    print("done.")


if __name__ == "__main__":
    main()
